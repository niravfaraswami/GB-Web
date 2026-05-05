#!/usr/bin/env python3
"""Generates an Excel workbook with a sheet literally named "Metafield Definitions"
so Matrixify recognises it as a definitions-import sheet.

Run: python3 build_definitions_xlsx.py
Produces: Sprout-Maker-Metafield-Definitions.xlsx
"""
import csv
from pathlib import Path
import openpyxl

SRC = Path(__file__).parent / "Sprout-Maker-Metafield-Definitions.csv"
OUT = Path(__file__).parent / "Sprout-Maker-Metafield-Definitions.xlsx"

with SRC.open(encoding="utf-8") as f:
    rows = list(csv.reader(f))

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Metafield Definitions"  # MUST match Matrixify's expected sheet name

for r in rows:
    ws.append(r)

# Reasonable column widths
widths = [10, 16, 14, 30, 30, 60, 30, 10, 22]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

wb.save(OUT)
print(f"Wrote {OUT}")
print(f"  Sheet: {ws.title}")
print(f"  Rows:  {ws.max_row}")
